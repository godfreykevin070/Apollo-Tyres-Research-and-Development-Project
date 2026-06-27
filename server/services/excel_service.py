import os
import json
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from config import config

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for handling Excel file operations"""
    
    def __init__(self):
        self.protocol_dir = config.PROTOCOL_DIR
        self.templates_dir = config.TEMPLATES_DIR
        
        # Ensure directories exist
        os.makedirs(self.protocol_dir, exist_ok=True)
    
    def read_excel_data(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Read Excel file and return data as list of dictionaries
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read (default: first sheet)
        
        Returns:
            List of dictionaries with column headers as keys
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        data = []
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return data
        
        # Get headers from first row
        headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(rows[0])]
        
        # Process data rows
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip empty rows
            
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    value = cell if cell is not None else ''
                    # Convert to string if needed
                    if isinstance(value, (int, float)):
                        value = str(value)
                    row_data[key] = value
            
            if any(row_data.values()):  # Only add if there's at least one non-empty value
                data.append(row_data)
        
        workbook.close()
        return data
    
    def write_excel_data(self, file_path: str, data: List[Dict[str, Any]], sheet_name: str = "Sheet1"):
        """
        Write data to Excel file
        
        Args:
            file_path: Path to save Excel file
            data: List of dictionaries to write
            sheet_name: Name of sheet
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        if not data:
            workbook.save(file_path)
            return
        
        # Get headers
        headers = list(data[0].keys())
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = row.get(key, '')
                cell.value = value
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_idx in range(2, len(data) + 2):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        workbook.save(file_path)
    
    def process_protocol_excel(self, file_path: str, replacements: Dict[str, str]) -> str:
        """
        Process a protocol Excel file by replacing placeholders
        
        Args:
            file_path: Path to the Excel file
            replacements: Dictionary of placeholder to value mappings
        
        Returns:
            Path to the processed file
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Protocol file not found: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for placeholder, value in replacements.items():
                            if placeholder in cell.value:
                                cell.value = cell.value.replace(placeholder, str(value) if value else '')
        
        # Save with output filename
        output_path = os.path.join(self.protocol_dir, 'output.xlsx')
        workbook.save(output_path)
        workbook.close()
        
        return output_path
    
    def extract_data_from_excel(self, file_path: str, header_row: int = 0) -> List[Dict[str, Any]]:
        """
        Extract data from Excel with custom header row
        
        Args:
            file_path: Path to Excel file
            header_row: Index of header row (0-based)
        
        Returns:
            List of dictionaries with column headers as keys
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        rows = list(sheet.iter_rows(values_only=True))
        
        if len(rows) <= header_row:
            workbook.close()
            return []
        
        # Get headers
        headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(rows[header_row])]
        
        data = []
        for row in rows[header_row + 1:]:
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    # Clean up key
                    key = key.replace(' ', '_').replace('[', '').replace(']', '').replace('(', '').replace(')', '')
                    key = key.lower()
                    value = cell if cell is not None else ''
                    row_data[key] = value
            
            if any(row_data.values()):
                data.append(row_data)
        
        workbook.close()
        return data
    
    def create_protocol_template(self, protocol: str, data: Dict[str, Any]) -> str:
        """
        Create a protocol Excel file from template with given data
        
        Args:
            protocol: Protocol name (MF62, MF52, etc.)
            data: Data to fill in the template
        
        Returns:
            Path to the created file
        """
        template_path = os.path.join(self.templates_dir, 'protocols', f"{protocol}.xlsx")
        
        if not os.path.exists(template_path):
            # Create a minimal template if not exists
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Write headers based on protocol
            if protocol == 'MF62':
                headers = ['Number Of Tests', 'Tests', 'Inflation Pressure [PSI]', 'Loads[Kg]', 
                          'Inclination Angle[°]', 'Slip Angle[°]', 'Slip Ratio [%]', 'Test Velocity [Kmph]',
                          'Job', 'Old Job', 'Template Tydex', 'Tydex name', 'P', 'L']
            elif protocol == 'MF52':
                headers = ['Number Of Tests', 'Tests', 'Inflation Pressure [PSI]', 'Loads[Kg]',
                          'Inclination Angle[°]', 'Slip Angle[°]', 'Slip Ratio [%]', 'Test Velocity [Kmph]',
                          'Job', 'Old Job', 'Template Tydex', 'Tydex name', 'P', 'L']
            elif protocol == 'FTire':
                headers = ['S.No', 'Test', 'Load (N)', 'IP (Kpa)', 'Speed (kmph)',
                          'Longitudinal Slip (%)', 'Slip Angle (deg)', 'Inclination Angle (deg)',
                          'Cleat Orientation (deg)', 'Job', 'Old Job', 'Template Tydex', 'Tydex name', 'P', 'L']
            elif protocol == 'CDTire':
                headers = ['No of Tests', 'Test Name', 'Inflation Pressure [bar]', 'Velocity [km/h]',
                          'Preload [N]', 'Camber [Deg]', 'Slip Angle [deg]', 'Displacement [mm]',
                          'Slip range [%]', 'Cleat', 'Road Surface', 'Job', 'Old Job', 
                          'Template Tydex', 'Tydex name', 'P', 'L']
            else:  # Custom
                headers = ['No of Tests', 'Tests', 'Inflation Pressure [PSI]', 'Loads [Kg]',
                          'Inclination Angle [°]', 'Slip Angle [°]', 'Slip Ratio [%]', 'Test Velocity [Kmph]',
                          'Cleat Orientation [°]', 'Displacement [mm]', 'Job', 'Old Job', 
                          'Template Tydex', 'Tydex name', 'P', 'L']
            
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            workbook.save(template_path)
        
        return template_path